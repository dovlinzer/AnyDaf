#!/usr/bin/env python3
"""
Generate taxonomy entries for all Tannaim and Amoraim.
Sources: Glida chronological table (PDF), Wikipedia (Hebrew) Tannaim/Amoraim pages.

Output: authorities_taxonomy.json + authorities_taxonomy.xlsx
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ---------------------------------------------------------------------------
# Data — compiled from Glida table + Wikipedia
# ---------------------------------------------------------------------------

# Each entry: (hebrew, english_name, aliases)
# Aliases include common alternate spellings/transliterations

PRE_TANNAITIC = [
    # Zugot and pre-Zugot
    ("שִׁמְעוֹן הַצַּדִּיק", "Shimon HaTzaddik", ["Simon the Just", "Simeon the Just"]),
    ("אַנְטִיגְנוֹס אִישׁ סוֹכוֹ", "Antignus of Socho", ["Antignos ish Socho"]),
    ("יוֹסֵי בֶּן יוֹעֶזֶר", "Yosi ben Yoezer", ["Yose ben Yoezer", "Jose ben Yoezer"]),
    ("יוֹסֵי בֶּן יוֹחָנָן", "Yosi ben Yochanan", ["Yose ben Yochanan"]),
    ("יְהוֹשֻׁעַ בֶּן פְּרַחְיָה", "Yehoshua ben Perachiah", ["Joshua ben Perachiah"]),
    ("נִתַּאי הָאַרְבֵּלִי", "Nittai HaArbelite", ["Nittai of Arbel"]),
    ("יְהוּדָה בֶּן טַבַּאי", "Yehudah ben Tabbai", ["Judah ben Tabbai"]),
    ("שִׁמְעוֹן בֶּן שֶׁטַח", "Shimon ben Shetach", ["Simeon ben Shetach"]),
    ("שְׁמַעְיָה", "Shemaiah", ["Shemaya"]),
    ("אַבְטַלְיוֹן", "Avtalion", ["Abtalion", "Avtalyon"]),
    ("הִלֵּל הַזָּקֵן", "Hillel the Elder", ["Hillel HaZaken", "Hillel"]),
    ("שַׁמַּאי הַזָּקֵן", "Shammai the Elder", ["Shammai HaZaken", "Shammai"]),
    ("חוֹנִי הַמְּעַגֵּל", "Honi HaMeagel", ["Honi the Circle-Drawer", "Honi HaMa'agal"]),
    ("עַקַבְיָא בֶּן מַהֲלַלְאֵל", "Akavya ben Mahalalel", ["Akavia ben Mahalalel"]),
    ("יוֹנָתָן בֶּן עֻזִּיאֵל", "Yonatan ben Uziel", ["Jonathan ben Uzziel"]),
    ("בָּבָא בֶּן בּוּטָא", "Bava ben Buta", ["Baba ben Buta"]),
    ("רַבָּן גַּמְלִיאֵל הַזָּקֵן", "Rabban Gamliel the Elder", ["Rabban Gamliel I"]),
    ("בֶּן בַּג בַּג", "Ben Bag Bag", []),
    ("נְחוּנְיָא בֶּן הַקָּנֶה", "Nechunia ben HaKanah", ["Nehuniah ben HaKanah"]),
]

TANNAIM_GEN1 = [
    # First generation — witnesses to the Destruction (~70-90 CE)
    ("רַבָּן יוֹחָנָן בֶּן זַכַּאי", "Rabban Yochanan ben Zakkai", ["Rabban Yohanan ben Zakkai"]),
    ("רַבָּן שִׁמְעוֹן בֶּן גַּמְלִיאֵל הַזָּקֵן", "Rabban Shimon ben Gamliel I", ["Rabban Simeon ben Gamliel I"]),
    ("רַבִּי חֲנִינָא סְגַן הַכֹּהֲנִים", "Rabbi Chanina, Deputy High Priest", ["Rabbi Chanina Segan HaKohanim"]),
    ("רַבִּי אֱלִיעֶזֶר בֶּן הוּרְקָנוֹס", "Rabbi Eliezer ben Horkeanus", ["Rabbi Eliezer ben Hyrcanus", "Rabbi Eliezer the Great"]),
    ("רַבִּי יְהוֹשֻׁעַ בֶּן חֲנַנְיָה", "Rabbi Yehoshua ben Chanania", ["Rabbi Joshua ben Chananya"]),
    ("רַבִּי אֶלְעָזָר בֶּן עֲרָךְ", "Rabbi Elazar ben Arach", []),
    ("רַבִּי יוֹסֵי הַכֹּהֵן", "Rabbi Yosi the Kohen", []),
    ("רַבִּי שִׁמְעוֹן בֶּן נְתַנְאֵל", "Rabbi Shimon ben Netanel", []),
    ("רַבִּי חֲנִינָא בֶּן דּוֹסָא", "Rabbi Chanina ben Dosa", ["Rabbi Hanina ben Dosa"]),
    ("רַבִּי דּוֹסָא בֶּן הַרְכִּינַס", "Rabbi Dosa ben Harkinas", []),
    ("רַבִּי אֱלִיעֶזֶר בֶּן יַעֲקֹב", "Rabbi Eliezer ben Yaakov", ["Rabbi Eliezer ben Jacob"]),
]

TANNAIM_GEN2 = [
    # Second generation — Yavneh period (~90-130 CE)
    ("רַבָּן גַּמְלִיאֵל דְּיַבְנֶה", "Rabban Gamliel of Yavneh", ["Rabban Gamliel II"]),
    ("רַבִּי אֶלְעָזָר בֶּן עֲזַרְיָה", "Rabbi Elazar ben Azariah", ["Rabbi Eleazar ben Azariah"]),
    ("רַבִּי טַרְפוֹן", "Rabbi Tarfon", ["Rabbi Tryphon"]),
    ("רַבִּי יוֹסֵי הַגְּלִילִי", "Rabbi Yosi the Galilean", ["Rabbi Jose the Galilean", "Rabbi Yose HaGalili"]),
    ("רַבִּי חֲלַפְתָּא", "Rabbi Chalafta", ["Rabbi Halafta of Tzippori"]),
    ("נַחוּם גַּמְזוֹ", "Nachum of Gamzo", ["Nahum of Gimzo"]),
    ("רַבִּי חֲנַנְיָה בֶּן תְּרַדְיוֹן", "Rabbi Chanania ben Teradyon", ["Rabbi Hanania ben Teradion"]),
    ("רַבִּי אֱלִיעֶזֶר בֶּן חִסְמָא", "Rabbi Elazar Chisma", ["Rabbi Eleazar Hisma"]),
    ("רַבִּי יוֹחָנָן בֶּן נוּרִי", "Rabbi Yochanan ben Nuri", ["Rabbi Yohanan ben Nuri"]),
]

TANNAIM_GEN3 = [
    # Third generation — Bar Kokhba era (~130-160 CE)
    ("רַבִּי עֲקִיבָא", "Rabbi Akiva", ["Rabbi Akiba", "Rabbi Aqiva"]),
    ("רַבִּי יִשְׁמָעֵאל", "Rabbi Yishmael", ["Rabbi Ishmael"]),
    ("שִׁמְעוֹן בֶּן עַזַּאי", "Shimon ben Azzai", ["Ben Azzai", "Simon ben Azzai"]),
    ("שִׁמְעוֹן בֶּן זוֹמָא", "Shimon ben Zoma", ["Ben Zoma"]),
    ("אֱלִישָׁע בֶּן אֲבוּיָה", "Elisha ben Avuyah", ["Acher", "Elisha ben Abuyah"]),
    ("רַבִּי יְהוּדָה בֶּן בָּבָא", "Rabbi Yehudah ben Bava", ["Rabbi Judah ben Bava"]),
    ("רַבִּי יוֹחָנָן בֶּן בְּרוֹקָא", "Rabbi Yochanan ben Beroka", []),
    ("רַבִּי יְהוּדָה בֶּן בְּתֵירָא", "Rabbi Yehudah ben Beteira", ["Rabbi Judah ben Batyra"]),
]

TANNAIM_GEN4 = [
    # Fourth generation — students of Akiva (~160-190 CE)
    ("רַבִּי מֵאִיר", "Rabbi Meir", []),
    ("רַבִּי יְהוּדָה בַּר אִלָּאי", "Rabbi Yehudah bar Ilai", ["Rabbi Judah bar Ilai", "Rabbi Yehudah"]),
    ("רַבִּי יוֹסֵי בֶּן חֲלַפְתָּא", "Rabbi Yosi ben Chalafta", ["Rabbi Jose ben Chalafta", "Rabbi Yose"]),
    ("רַבִּי שִׁמְעוֹן בַּר יוֹחַאי", "Rabbi Shimon bar Yochai", ["Rashbi", "Rabbi Simeon bar Yochai"]),
    ("רַבִּי אֶלְעָזָר בֶּן שַׁמּוּעַ", "Rabbi Elazar ben Shamua", ["Rabbi Eleazar ben Shamua"]),
    ("רַבִּי יְהוֹשֻׁעַ בֶּן קָרְחָה", "Rabbi Yehoshua ben Karchah", ["Rabbi Joshua ben Karhah"]),
    ("רַבִּי נָתָן הַבַּבְלִי", "Rabbi Natan the Babylonian", ["Rabbi Nathan the Babylonian"]),
    ("רַבִּי יוֹחָנָן הַסַּנְדְּלָר", "Rabbi Yochanan the Sandal-Maker", ["Rabbi Yohanan the Sandal-Maker"]),
    ("בְּרוּרְיָה", "Beruriah", ["Beruria"]),
    ("רַבִּי שִׁמְעוֹן בֶּן אֶלְעָזָר", "Rabbi Shimon ben Elazar", []),
    ("רַבִּי שִׁמְעוֹן בֶּן מֶנַסְיָא", "Rabbi Shimon ben Manasia", []),
    ("רַבִּי פִּינְחָס בֶּן יָאִיר", "Rabbi Pinchas ben Yair", ["Rabbi Phinehas ben Jair"]),
]

TANNAIM_GEN5 = [
    # Fifth generation — Mishna editors (~190-220 CE)
    ("רַבִּי יְהוּדָה הַנָּשִׂיא", "Rabbi Yehudah HaNasi", ["Rebbi", "Rabbi Judah the Prince", "Rebbi HaNasi"]),
    ("רַבָּן שִׁמְעוֹן בֶּן גַּמְלִיאֵל", "Rabban Shimon ben Gamliel II", []),
    ("רַבִּי יִשְׁמָעֵאל בַּר רַבִּי יוֹסֵי", "Rabbi Yishmael bar Rabbi Yosi", []),
    ("רַבִּי יוֹסֵי בַּר יְהוּדָה", "Rabbi Yosi bar Yehudah", []),
    ("רַבִּי אֶלְעָזָר בַּר רַבִּי שִׁמְעוֹן", "Rabbi Elazar bar Rabbi Shimon", []),
    ("סֻמְכוּס", "Sumchus", ["Symmachus"]),
    ("רַבִּי אֶלְעָזָר הַקַּפָּר", "Rabbi Elazar HaKapar", ["Rabbi Eleazar Hakappar", "Bar Kappara's father"]),
    ("אִסִּי בֶּן יְהוּדָה", "Issi ben Yehudah", []),
]

TRANSITIONAL = [
    # Bridge generation between Tannaim and Amoraim (~210-230 CE)
    ("רַבִּי חִיָּיא", "Rabbi Chiyya", ["Rabbi Hiyya", "Rabbi Hiya"]),
    ("בַּר קַפָּרָא", "Bar Kappara", ["Bar Kappara"]),
    ("רַבִּי אוֹשַׁעְיָה רַבָּה", "Rabbi Hoshaia Rabbah", ["Rabbi Oshaya Rabbah", "Rabbi Oshaiah"]),
    ("רַבִּי אַפֶס", "Rabbi Afes", ["Rabbi Efes"]),
    ("לֵוִי", "Levi (Transitional)", ["Levi bar Sisi"]),
]

AMORAIM_EY_GEN1 = [
    # Eretz Yisrael, 1st generation (~220-250 CE)
    ("רַבִּי יְהוֹשֻׁעַ בֶּן לֵוִי", "Rabbi Yehoshua ben Levi", ["Rabbi Joshua ben Levi"]),
    ("רַבִּי יַנַּאי", "Rabbi Yannai", ["Rabbi Yannai of Akbara"]),
    ("רַבִּי חֲנִינָא", "Rabbi Chanina (bar Chama)", ["Rabbi Hanina bar Hama"]),
    ("רַבִּי יְהוּדָה נְשִׂיאָה א'", "Rabbi Yehudah Nesi'ah I", ["Rabbi Judah Nesi'ah I"]),
    ("רַבִּי יוֹנָתָן בֶּן אֶלְעָזָר", "Rabbi Yonatan ben Elazar", []),
]

AMORAIM_EY_GEN2 = [
    # Eretz Yisrael, 2nd generation (~250-280 CE)
    ("רַבִּי יוֹחָנָן בַּר נַפְחָא", "Rabbi Yochanan bar Nafcha", ["Rabbi Yohanan bar Napaha", "Rabbi Yochanan"]),
    ("רֵישׁ לָקִישׁ", "Reish Lakish", ["Rabbi Shimon ben Lakish", "Resh Lakish"]),
    ("רַבִּי יוֹסֵי בַּר חֲנִינָא", "Rabbi Yosi bar Chanina", ["Rabbi Jose bar Hanina"]),
    ("רַבִּי שְׁמוּאֵל בַּר נַחְמָנִי", "Rabbi Shmuel bar Nachmani", ["Rabbi Samuel bar Nahmani"]),
    ("רַבִּי שִׂמְלַאי", "Rabbi Simlai", ["Rabbi Shmuelai of Lod"]),
    ("רַבִּי לֵוִי", "Rabbi Levi", []),
]

AMORAIM_EY_GEN3 = [
    # Eretz Yisrael, 3rd generation (~280-310 CE)
    ("רַבִּי אֶלְעָזָר בֶּן פְּדָת", "Rabbi Elazar ben Pedat", ["Rabbi Eleazar ben Pedat"]),
    ("רַבִּי אַמִּי", "Rabbi Ami", ["Rabbi Ammi"]),
    ("רַבִּי אַסִּי", "Rabbi Assi", ["Rabbi Assi of Tiberias"]),
    ("רַבִּי אַבָּהוּ", "Rabbi Abbahu", ["Rabbi Abbahu of Caesarea"]),
    ("רַבִּי זֵירָא", "Rabbi Zeira", ["Rabbi Zera", "Rav Zeira"]),
    ("רַבִּי חִיָּיא בַּר אַבָּא", "Rabbi Chiyya bar Abba", ["Rabbi Hiyya bar Abba"]),
    ("יִצְחָק נַפְחָא", "Rabbi Yitzchak Nafcha", ["Rabbi Isaac Napcha"]),
    ("רַבִּי אִלְעַאי", "Rabbi Ila'i", ["Rabbi Illai of Tzor"]),
    ("רַבִּי חֶלְבּוֹ", "Rabbi Chelbo", ["Rabbi Helbo"]),
    ("רַבִּי אַבָּא", "Rabbi Abba (bar Zavda)", []),
    ("רַבִּי יַעֲקֹב בַּר אִידִי", "Rabbi Yaakov bar Idi", []),
    ("רַבִּי שִׁמְעוֹן בֶּן פָּזִי", "Rabbi Shimon ben Pazi", []),
    ("רַבִּי יְהוּדָה נְשִׂיאָה ב'", "Rabbi Yehudah Nesi'ah II", []),
    ("רַבָּה בַּר בַּר חָנָה", "Rabba bar bar Chana", ["Rabbah bar bar Hannah"]),
    ("עוּלָּא", "Ulla (bar Yishmael)", ["Ulla", "Rabbi Ulla"]),
]

AMORAIM_EY_GEN4 = [
    # Eretz Yisrael, 4th generation (~310-340 CE)
    ("רַבִּי יִרְמְיָה", "Rabbi Yirmiyah", ["Rabbi Jeremiah"]),
    ("רַבִּי יוֹנָה", "Rabbi Yona", ["Rabbi Jonah"]),
    ("רַבִּי יוֹסֵי בַּר זְבִידָא", "Rabbi Yosi bar Zavida", ["Rabbi Jose bar Zavida"]),
    ("רַבִּי בֶּרֶכְיָה", "Rabbi Berechia", ["Rabbi Berechiah"]),
    ("רַבִּי יוֹסֵי (טְבֶרְיָה)", "Rabbi Yosi of Tiberias", []),
    ("רַבָּן הִלֵּל (הַלּוּחַ)", "Rabban Hillel II", ["Hillel II", "Hillel the Calendar"]),
    ("רַבִּי אַחָא", "Rabbi Acha", ["Rabbi Aha"]),
    ("רַבִּי פְּדָת", "Rabbi Pedat", []),
]

AMORAIM_EY_GEN5 = [
    # Eretz Yisrael, 5th generation (~340-380 CE)
    ("רַבִּי מָנָא", "Rabbi Mana", ["Rabbi Mana II"]),
    ("רַבִּי יוֹסֵי בַּר אָבִין", "Rabbi Yosi bar Avin", ["Rabbi Jose bar Bun", "Rabbi Yosi bar Bun"]),
    ("רַבִּי אָבִין", "Rabbi Avin", ["Ravin", "Rabbi Abin"]),
    ("רַבִּי תַּנְחוּמָא", "Rabbi Tanchuma", ["Rabbi Tanhuma"]),
    ("רַבִּי יְהוּדָה בַּר שָׁלוֹם", "Rabbi Yehudah bar Shalom", []),
    ("רַבִּי חִזְקִיָּה", "Rabbi Chizkiah", ["Rabbi Hezekiah"]),
]

AMORAIM_EY_GEN6 = [
    # Eretz Yisrael, 6th generation (~380-410 CE)
    ("רַבִּי יוֹסֵי בַּר רַבִּי אָבוּן", "Rabbi Yosi bar Rabbi Abun", ["Rabbi Yose bar Rabbi Bun", "final redactor of Yerushalmi"]),
    ("רַבִּי חֲנִינָא דְּצִיפּוֹרִין", "Rabbi Chanina of Tzippori", []),
]

AMORAIM_BAB_GEN1 = [
    # Babylon, 1st generation (~220-250 CE)
    ("רַב (אַרִיכָא)", "Rav (Abba Arikha)", ["Rav", "Abba Arikha", "Rav Abba"]),
    ("שְׁמוּאֵל (נְהַרְדְּעָא)", "Shmuel of Nehardea", ["Mar Shmuel", "Samuel of Nehardea"]),
    ("רַב שֵׁילָא", "Rav Shila", ["Rav Sheila"]),
    ("רַב אַסִּי", "Rav Assi", ["Rav Asi"]),
    ("מַר עוּקְבָּא", "Mar Ukva", ["Mar Ukba"]),
    ("זְעֵירִי", "Ze'iri", ["Zeiri"]),
    ("רַבָּה בַּר חָנָה", "Rabba bar Chana", ["Rabbah bar Hana"]),
    ("רַב כַּהֲנָא (א')", "Rav Kahana I", ["Rav Kahana the First"]),
]

AMORAIM_BAB_GEN2 = [
    # Babylon, 2nd generation (~250-280 CE)
    ("רַב הוּנָא (סוּרָא)", "Rav Huna of Sura", ["Rav Huna"]),
    ("רַב יְהוּדָה בַּר יְחֶזְקֵאל", "Rav Yehudah bar Yechezkel", ["Rav Judah bar Ezekiel"]),
    ("רַב נַחְמָן (בַּר יַעֲקֹב)", "Rav Nachman bar Yaakov", ["Rav Nahman bar Jacob"]),
    ("רַב אָדָא בַּר אַהֲבָה", "Rav Ada bar Ahava", ["Rav Ada bar Ahavah"]),
    ("חִיָּיא בַּר רַב", "Chiyya bar Rav", ["Hiyya bar Rav"]),
    ("רַב גִּידֵּל", "Rav Giddel", ["Rav Gidal"]),
    ("רַב יִרְמְיָה בַּר אַבָּא", "Rav Yirmiyah bar Abba", ["Rav Jeremiah bar Abba"]),
    ("אָבִימִי", "Avimi", ["Avimi of Haifa"]),
    ("רַב כַּהֲנָא (ב')", "Rav Kahana II", []),
]

AMORAIM_BAB_GEN3 = [
    # Babylon, 3rd generation (~280-310 CE)
    ("רַבָּה בַּר נַחְמָנִי", "Rabba bar Nachmani", ["Rabbah bar Nahmani", "Rabbah"]),
    ("רַב יוֹסֵף בַּר חִיָּיא", "Rav Yosef bar Chiyya", ["Rav Joseph", "Rav Yosef"]),
    ("רַב חִסְדָּא", "Rav Chisda", ["Rav Hisda"]),
    ("רַב שֵׁשֶׁת", "Rav Sheshet", ["Rav Sheishet"]),
    ("רַב הַמְנוּנָא", "Rav Hamnuna", ["Rav Hamnona"]),
    ("רַב סָפְרָא", "Rav Safra", []),
    ("רַב זְבִיד (א')", "Rav Zavid I", []),
    ("רַב יִצְחָק בַּר אַבְדִּימִי", "Rav Yitzchak bar Avdimi", []),
]

AMORAIM_BAB_GEN4 = [
    # Babylon, 4th generation (~310-340 CE)
    ("אַבָּיֵי", "Abaye", ["Abaye", "Abayei"]),
    ("רָבָא", "Rava", ["Raba", "Rava bar Yosef"]),
    ("רָמִי בַּר חָמָא", "Rami bar Chama", ["Rami bar Hama"]),
    ("רַב נַחְמָן בַּר יִצְחָק", "Rav Nachman bar Yitzchak", ["Rav Nahman bar Isaac"]),
    ("רַב זְבִיד (מִנְהַרְדְּעָא)", "Rav Zavid of Nehardea", []),
    ("רַב אִידִי בַּר אָבִין", "Rav Idi bar Avin", []),
    ("רַב חָמָא", "Rav Chama", []),
    ("רַב דִּימִי", "Rav Dimi", ["Rav Dimi of Nehardea"]),
]

AMORAIM_BAB_GEN5 = [
    # Babylon, 5th generation (~340-380 CE)
    ("רַב פָּפָּא (נֶרֶשׁ)", "Rav Papa", ["Rav Pappa"]),
    ("רַב הוּנָא בְּרֵיהּ דְּרַב יְהוֹשֻׁעַ", "Rav Huna brei d'Rav Yehoshua", ["Rav Huna son of Rav Joshua"]),
    ("רַב כַּהֲנָא (ג')", "Rav Kahana III", []),
    ("רַב אַחָא בְּרֵיהּ דְּרָבָא", "Rav Acha brei d'Rava", []),
    ("מָרֵימָר", "Meremar", ["Maremar"]),
    ("רַב אִידִי בַּר אָבִין (ב')", "Rav Idi bar Avin II", []),
    ("רַפְרָם (א')", "Rafram I", []),
    ("רַב גְּבִיהָה מִבֵּי כָּתִיל", "Rav Geviha of Bei Katil", []),
]

AMORAIM_BAB_GEN6 = [
    # Babylon, 6th generation (~380-430 CE)
    ("רַב אָשִׁי", "Rav Ashi", ["Rav Ashi of Mata Mechasia"]),
    ("רְבִינָא (א')", "Ravina I", ["Ravina bar Huna", "Ravina the First"]),
    ("אָמֵימָר", "Ameimar", ["Amimemar"]),
    ("מַר זוּטְרָא", "Mar Zutra", ["Mar Zutra I"]),
    ("רַב אִידִי בַּר אָבִין (ב')", "Rav Idi bar Avin (later)", []),
    ("מָרֵימָר (ב')", "Meremar II", []),
    ("רַפְרָם (ב')", "Rafram II", []),
    ("רַב רְחוּמֵי", "Rav Ruchami", ["Rav Rehumi"]),
]

AMORAIM_BAB_GEN7 = [
    # Babylon, 7th generation / Savoraim transitional (~430-500 CE)
    ("מַר בַּר רַב אָשִׁי", "Mar bar Rav Ashi", ["Tavyomi"]),
    ("רָבִינָא (ב')", "Ravina II", ["Ravina bar Huna (last)", "Ravina the Last", "Chotemet HaTalmud"]),
    ("רַב יוֹסֵף (חוֹתֵם הַתַּלְמוּד)", "Rav Yosef II", ["Sealer of the Talmud"]),
    ("רַב סִימוֹנָא", "Rav Simona", []),
    ("רַב הַנָּן", "Rav Hanan", []),
]

# ---------------------------------------------------------------------------
# Build taxonomy entries
# ---------------------------------------------------------------------------

def make_id(english_name: str) -> str:
    s = english_name.lower()
    for c in ["'", "'", "(", ")", ",", ".", "'"]:
        s = s.replace(c, "")
    s = s.replace(" ", "_")
    # Shorten common prefixes
    s = s.replace("rabbi_", "r_").replace("rabban_", "rbn_").replace("rav_", "rv_")
    return "auth_" + s


def make_entries(sages, category, parent_id, period_label, geography):
    entries = []
    for hebrew, english, aliases in sages:
        eid = make_id(english)
        all_aliases = [english, hebrew] + aliases
        # Add geography/period to aliases for searchability
        entry = {
            "id": eid,
            "name": english,
            "hebrew": hebrew,
            "parent_id": parent_id,
            "category": category,
            "sources": ["authorities"],
            "notes": f"{period_label} | {geography}",
            "aliases": all_aliases,
        }
        entries.append(entry)
    return entries


# Build period group entries (the parent nodes)
PERIOD_GROUPS = [
    {
        "id": "auth_pre_tannaitic",
        "name": "Pre-Tannaitic Sages (Zugot and Earlier)",
        "hebrew": "חַכְמֵי הַזּוּגוֹת",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~300 BCE – 70 CE",
        "aliases": ["Zugot", "Pre-Tannaim", "זוגות"],
    },
    {
        "id": "auth_tannaim_gen1",
        "name": "Tannaim, First Generation",
        "hebrew": "תַּנָּאִים, דּוֹר רִאשׁוֹן",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~70-90 CE | Eretz Yisrael",
        "aliases": ["Tannaim first generation"],
    },
    {
        "id": "auth_tannaim_gen2",
        "name": "Tannaim, Second Generation",
        "hebrew": "תַּנָּאִים, דּוֹר שֵׁנִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~90-130 CE | Yavneh",
        "aliases": ["Tannaim second generation", "Yavneh sages"],
    },
    {
        "id": "auth_tannaim_gen3",
        "name": "Tannaim, Third Generation",
        "hebrew": "תַּנָּאִים, דּוֹר שְׁלִישִׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~130-160 CE | Bar Kokhba era",
        "aliases": ["Tannaim third generation"],
    },
    {
        "id": "auth_tannaim_gen4",
        "name": "Tannaim, Fourth Generation",
        "hebrew": "תַּנָּאִים, דּוֹר רְבִיעִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~160-190 CE | Students of Akiva",
        "aliases": ["Tannaim fourth generation", "students of Akiva"],
    },
    {
        "id": "auth_tannaim_gen5",
        "name": "Tannaim, Fifth Generation",
        "hebrew": "תַּנָּאִים, דּוֹר חֲמִישִׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~190-220 CE | Mishna editors",
        "aliases": ["Tannaim fifth generation", "Mishna editors"],
    },
    {
        "id": "auth_transitional",
        "name": "Transitional Generation (Tanna-Amora Bridge)",
        "hebrew": "דּוֹר הַמַּעֲבָר",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — TANNAIM",
        "sources": ["authorities"],
        "notes": "~210-230 CE",
        "aliases": ["Dor HaMa'avar", "transitional generation"],
    },
    {
        "id": "auth_amora_ey_gen1",
        "name": "Palestinian Amoraim, First Generation",
        "hebrew": "אָמוֹרָאֵי אֶרֶץ יִשְׂרָאֵל, דּוֹר רִאשׁוֹן",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~220-250 CE | Eretz Yisrael",
        "aliases": ["Palestinian Amoraim first generation", "Amoraim EY gen 1"],
    },
    {
        "id": "auth_amora_ey_gen2",
        "name": "Palestinian Amoraim, Second Generation",
        "hebrew": "אָמוֹרָאֵי אֶרֶץ יִשְׂרָאֵל, דּוֹר שֵׁנִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~250-280 CE | Eretz Yisrael",
        "aliases": [],
    },
    {
        "id": "auth_amora_ey_gen3",
        "name": "Palestinian Amoraim, Third Generation",
        "hebrew": "אָמוֹרָאֵי אֶרֶץ יִשְׂרָאֵל, דּוֹר שְׁלִישִׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~280-310 CE | Eretz Yisrael",
        "aliases": [],
    },
    {
        "id": "auth_amora_ey_gen4",
        "name": "Palestinian Amoraim, Fourth Generation",
        "hebrew": "אָמוֹרָאֵי אֶרֶץ יִשְׂרָאֵל, דּוֹר רְבִיעִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~310-340 CE | Eretz Yisrael",
        "aliases": [],
    },
    {
        "id": "auth_amora_ey_gen5",
        "name": "Palestinian Amoraim, Fifth Generation",
        "hebrew": "אָמוֹרָאֵי אֶרֶץ יִשְׂרָאֵל, דּוֹר חֲמִישִׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~340-380 CE | Eretz Yisrael",
        "aliases": [],
    },
    {
        "id": "auth_amora_ey_gen6",
        "name": "Palestinian Amoraim, Sixth Generation",
        "hebrew": "אָמוֹרָאֵי אֶרֶץ יִשְׂרָאֵל, דּוֹר שִׁשִּׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~380-410 CE | Eretz Yisrael",
        "aliases": [],
    },
    {
        "id": "auth_amora_bab_gen1",
        "name": "Babylonian Amoraim, First Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר רִאשׁוֹן",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~220-250 CE | Babylon",
        "aliases": [],
    },
    {
        "id": "auth_amora_bab_gen2",
        "name": "Babylonian Amoraim, Second Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר שֵׁנִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~250-280 CE | Babylon",
        "aliases": [],
    },
    {
        "id": "auth_amora_bab_gen3",
        "name": "Babylonian Amoraim, Third Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר שְׁלִישִׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~280-310 CE | Babylon",
        "aliases": [],
    },
    {
        "id": "auth_amora_bab_gen4",
        "name": "Babylonian Amoraim, Fourth Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר רְבִיעִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~310-340 CE | Babylon (Abaye and Rava generation)",
        "aliases": ["Abaye and Rava generation"],
    },
    {
        "id": "auth_amora_bab_gen5",
        "name": "Babylonian Amoraim, Fifth Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר חֲמִישִׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~340-380 CE | Babylon",
        "aliases": [],
    },
    {
        "id": "auth_amora_bab_gen6",
        "name": "Babylonian Amoraim, Sixth Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר שִׁשִּׁי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~380-430 CE | Babylon (Rav Ashi generation)",
        "aliases": ["Rav Ashi generation"],
    },
    {
        "id": "auth_amora_bab_gen7",
        "name": "Babylonian Amoraim, Seventh Generation",
        "hebrew": "אָמוֹרָאֵי בָּבֶל, דּוֹר שְׁבִיעִי",
        "parent_id": None,
        "category": "RABBINIC AUTHORITIES — AMORAIM",
        "sources": ["authorities"],
        "notes": "~430-500 CE | Babylon (Sealers of the Talmud)",
        "aliases": ["Ravina and Rav Ashi", "Chotemet HaTalmud"],
    },
]

# Build sage entries for each period
SAGE_GROUPS = [
    (PRE_TANNAITIC,    "RABBINIC AUTHORITIES — TANNAIM", "auth_pre_tannaitic",  "Pre-Tannaitic (~300 BCE–70 CE)",            "Eretz Yisrael"),
    (TANNAIM_GEN1,     "RABBINIC AUTHORITIES — TANNAIM", "auth_tannaim_gen1",   "Tannaim, First Generation (~70–90 CE)",      "Eretz Yisrael"),
    (TANNAIM_GEN2,     "RABBINIC AUTHORITIES — TANNAIM", "auth_tannaim_gen2",   "Tannaim, Second Generation (~90–130 CE)",    "Eretz Yisrael"),
    (TANNAIM_GEN3,     "RABBINIC AUTHORITIES — TANNAIM", "auth_tannaim_gen3",   "Tannaim, Third Generation (~130–160 CE)",    "Eretz Yisrael"),
    (TANNAIM_GEN4,     "RABBINIC AUTHORITIES — TANNAIM", "auth_tannaim_gen4",   "Tannaim, Fourth Generation (~160–190 CE)",   "Eretz Yisrael"),
    (TANNAIM_GEN5,     "RABBINIC AUTHORITIES — TANNAIM", "auth_tannaim_gen5",   "Tannaim, Fifth Generation (~190–220 CE)",    "Eretz Yisrael"),
    (TRANSITIONAL,     "RABBINIC AUTHORITIES — TANNAIM", "auth_transitional",   "Transitional Generation (~210–230 CE)",      "Both"),
    (AMORAIM_EY_GEN1,  "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_ey_gen1",  "Palestinian Amoraim, 1st Gen (~220–250 CE)", "Eretz Yisrael"),
    (AMORAIM_EY_GEN2,  "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_ey_gen2",  "Palestinian Amoraim, 2nd Gen (~250–280 CE)", "Eretz Yisrael"),
    (AMORAIM_EY_GEN3,  "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_ey_gen3",  "Palestinian Amoraim, 3rd Gen (~280–310 CE)", "Eretz Yisrael"),
    (AMORAIM_EY_GEN4,  "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_ey_gen4",  "Palestinian Amoraim, 4th Gen (~310–340 CE)", "Eretz Yisrael"),
    (AMORAIM_EY_GEN5,  "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_ey_gen5",  "Palestinian Amoraim, 5th Gen (~340–380 CE)", "Eretz Yisrael"),
    (AMORAIM_EY_GEN6,  "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_ey_gen6",  "Palestinian Amoraim, 6th Gen (~380–410 CE)", "Eretz Yisrael"),
    (AMORAIM_BAB_GEN1, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen1", "Babylonian Amoraim, 1st Gen (~220–250 CE)",  "Babylon"),
    (AMORAIM_BAB_GEN2, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen2", "Babylonian Amoraim, 2nd Gen (~250–280 CE)",  "Babylon"),
    (AMORAIM_BAB_GEN3, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen3", "Babylonian Amoraim, 3rd Gen (~280–310 CE)",  "Babylon"),
    (AMORAIM_BAB_GEN4, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen4", "Babylonian Amoraim, 4th Gen (~310–340 CE)",  "Babylon"),
    (AMORAIM_BAB_GEN5, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen5", "Babylonian Amoraim, 5th Gen (~340–380 CE)",  "Babylon"),
    (AMORAIM_BAB_GEN6, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen6", "Babylonian Amoraim, 6th Gen (~380–430 CE)",  "Babylon"),
    (AMORAIM_BAB_GEN7, "RABBINIC AUTHORITIES — AMORAIM", "auth_amora_bab_gen7", "Babylonian Amoraim, 7th Gen (~430–500 CE)",  "Babylon"),
]

all_entries = list(PERIOD_GROUPS)
for sages, cat, parent_id, period_label, geo in SAGE_GROUPS:
    all_entries.extend(make_entries(sages, cat, parent_id, period_label, geo))

# ---------------------------------------------------------------------------
# Save JSON
# ---------------------------------------------------------------------------

out_dir = Path(__file__).parent
json_path = out_dir / "authorities_taxonomy.json"
json_path.write_text(json.dumps({"entries": all_entries}, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {json_path} — {len(all_entries)} entries")

# ---------------------------------------------------------------------------
# Save Excel
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Authorities"

headers = ["Category", "Entry ID", "Name (English)", "Hebrew", "Parent Group", "Period | Geography", "Aliases"]
ws.append(headers)

hdr_fill = PatternFill(start_color="1B3A8A", end_color="1B3A8A", fill_type="solid")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center")

tanna_fill  = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
amora_ey_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
amora_bab_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
group_font = Font(bold=True)

parent_names = {e["id"]: e["name"] for e in all_entries}

for row_num, e in enumerate(all_entries, start=2):
    cat = e.get("category", "")
    if "TANNAIM" in cat:
        fill = tanna_fill
    elif "AMORAIM" in cat and "EY" in e["id"]:
        fill = amora_ey_fill
    else:
        fill = amora_bab_fill

    parent_name = parent_names.get(e.get("parent_id", ""), "")
    aliases_str = " | ".join(a for a in (e.get("aliases") or []) if a != e["name"] and a != e.get("hebrew"))
    row = [
        cat,
        e["id"],
        e["name"],
        e.get("hebrew", ""),
        parent_name,
        e.get("notes", ""),
        aliases_str,
    ]
    ws.append(row)
    is_group = not e.get("parent_id")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=(col in [3, 4, 7]))
        if is_group:
            cell.font = group_font

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 28
ws.column_dimensions["E"].width = 38
ws.column_dimensions["F"].width = 32
ws.column_dimensions["G"].width = 45
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(all_entries)+1}"

xlsx_path = out_dir / "authorities_taxonomy.xlsx"
wb.save(xlsx_path)
print(f"Wrote {xlsx_path}")
print(f"\nBreakdown:")
print(f"  Period group headers: {sum(1 for e in all_entries if not e.get('parent_id'))}")
print(f"  Individual sages:     {sum(1 for e in all_entries if e.get('parent_id'))}")
