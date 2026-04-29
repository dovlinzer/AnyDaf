#!/usr/bin/env python3
"""Merge authorities_taxonomy into seed_taxonomy, replacing flat tannaim/amoraim entries."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DIR = Path(__file__).parent

# IDs to remove from seed taxonomy (flat, incomplete tannaim/amoraim)
REMOVE_IDS = {
    # top-level authority group nodes (replaced by authorities_taxonomy period groups)
    "tannaim", "amoraim",
    # flat tannaim entries
    "tanna_hillel", "tanna_shammai", "tanna_beit_hillel", "tanna_beit_shammai",
    "tanna_raban_gamliel", "tanna_rybz", "tanna_rabbi_eliezer", "tanna_rabbi_yehoshua",
    "tanna_rabbi_akiva", "tanna_rabbi_yishmael", "tanna_rabbi_tarfon", "tanna_rabbi_meir",
    "tanna_rabbi_yehudah_bar_ilai", "tanna_rabbi_shimon_bar_yochai", "tanna_rabbi_yosi_bar_chalafta",
    "tanna_rabbi_elazar_ben_azaryah", "tanna_ben_azzai", "tanna_ben_zoma",
    "tanna_rebbi", "tanna_bar_kappara", "tanna_rebbe_shimon_ben_gamliel",
    # flat amoraim group nodes and entries
    "amora_palestinian", "amora_babylonian",
    "amora_rabbi_yochanan", "amora_reish_lakish", "amora_rabbi_chanina",
    "amora_rabbi_oshiya", "amora_rabbi_yannai", "amora_rabbi_yehoshua_ben_levi",
    "amora_rabbi_elazar_ben_pedat", "amora_rabbi_ami", "amora_rabbi_asi",
    "amora_rabbi_abbahu", "amora_rav_zeira", "amora_rabbi_jeremiah",
    "amora_rabbi_mana", "amora_rav", "amora_shmuel", "amora_rav_huna",
    "amora_rav_yehudah", "amora_rav_chisda", "amora_rav_sheshet",
    "amora_rav_nachman", "amora_rabba", "amora_rav_yosef",
    "amora_rav_nachman_bar_yitzchak", "amora_abaye", "amora_rava",
    "amora_rav_papa", "amora_rav_ashi", "amora_ravina",
}

# Load seed taxonomy
seed_data = json.loads((DIR / "seed_taxonomy.json").read_text(encoding="utf-8"))
seed_entries = seed_data if isinstance(seed_data, list) else seed_data.get("entries", [])

# Load authorities taxonomy
auth_data = json.loads((DIR / "authorities_taxonomy.json").read_text(encoding="utf-8"))
auth_entries = auth_data if isinstance(auth_data, list) else auth_data.get("entries", [])

# Filter seed: remove flat tannaim/amoraim entries
filtered_seed = [e for e in seed_entries if e["id"] not in REMOVE_IDS]
removed = len(seed_entries) - len(filtered_seed)

# Merge: seed (filtered) + authorities
merged = filtered_seed + auth_entries

print(f"Seed entries:       {len(seed_entries)}")
print(f"Removed (flat):     {removed}")
print(f"Authorities added:  {len(auth_entries)}")
print(f"Merged total:       {len(merged)}")

# Validate: no broken parent_id references
ids = {e["id"] for e in merged}
broken = [(e["id"], e["parent_id"]) for e in merged if e.get("parent_id") and e["parent_id"] not in ids]
if broken:
    print(f"\nWARNING — broken parent_id references: {len(broken)}")
    for child_id, parent_id in broken:
        print(f"  {child_id} → {parent_id} (missing)")
else:
    print("Parent ID validation: OK — no broken references")

# Save merged JSON
out_json = DIR / "merged_taxonomy.json"
out_json.write_text(json.dumps({"entries": merged}, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nWrote {out_json} — {len(merged)} entries")

# --- Build Excel ---
# Collect categories and assign colors
CATEGORY_COLORS = [
    "FFF2CC","FFE6CC","FFD9D9","FFD9F0","F0D9FF","D9E8FF","D9F5FF","D9FFE8",
    "E8FFD9","F5FFD9","FFFBD9","FFE8D9","F0F0F0","D9D9FF","FFD9D9","D9FFD9",
    "D9F0FF","FFD9FF","FFF0D9","D9FFFF","E0D9FF","FFD9E0","D9FFE0","E0FFD9",
    "D9E0FF","FFE0D9","D9FFD9","FFD9D9","D9D9D9","F9F9F9",
]

cats = []
seen_cats = set()
for e in merged:
    c = e.get("category", "Uncategorized")
    if c not in seen_cats:
        cats.append(c)
        seen_cats.add(c)
cat_color = {c: CATEGORY_COLORS[i % len(CATEGORY_COLORS)] for i, c in enumerate(cats)}

id_to_name = {e["id"]: e["name"] for e in merged}

wb = Workbook()
ws = wb.active
ws.title = "Merged Taxonomy"

# Header
headers = ["Category", "Entry ID", "Name", "Hebrew", "Parent Entry", "Aliases", "Notes"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="404040")
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 38
ws.column_dimensions["F"].width = 42
ws.column_dimensions["G"].width = 38
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G1"

row = 2
for e in merged:
    cat = e.get("category", "Uncategorized")
    color = cat_color.get(cat, "FFFFFF")
    is_top = not e.get("parent_id")
    parent_name = id_to_name.get(e.get("parent_id", ""), "")
    aliases = ", ".join(e.get("aliases", []))

    vals = [cat, e["id"], e["name"], e.get("hebrew", ""), parent_name, aliases, e.get("notes", "")]
    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=color)
        if is_top:
            cell.font = Font(bold=True)
        if col == 4:
            cell.alignment = Alignment(horizontal="right")
    row += 1

out_xlsx = DIR / "merged_taxonomy.xlsx"
wb.save(out_xlsx)
print(f"Wrote {out_xlsx}")
