#!/usr/bin/env python3
"""
mishnah_gap_fill.py — Extract missing taxonomy entries from the Mishnah via Sefaria.

For each tractate, sends the English Mishnah text + current taxonomy to Claude and asks
it to identify halakhic/conceptual topics discussed in that tractate that are not yet
represented in the taxonomy.

Usage:
  python mishnah_gap_fill.py --seder Moed          # all tractates in Seder Moed
  python mishnah_gap_fill.py --tractate Shabbat     # single tractate
  python mishnah_gap_fill.py --all                  # all 6 sedarim (slow)
  python mishnah_gap_fill.py --merge                # combine all tractate outputs → candidates.json + xlsx
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.request
import urllib.parse
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv(Path(__file__).parent.parent.parent / ".env", override=True)

DIR = Path(__file__).parent

SEDARIM = {
    "Zeraim":   ["Berakhot", "Peah", "Demai", "Kilayim", "Sheviit", "Terumot",
                 "Maasrot", "Maaser Sheni", "Challah", "Orlah", "Bikkurim"],
    "Moed":     ["Shabbat", "Eruvin", "Pesachim", "Shekalim", "Yoma", "Sukkah",
                 "Beitzah", "Rosh Hashanah", "Taanit", "Megillah", "Moed Katan", "Chagigah"],
    "Nashim":   ["Yevamot", "Ketubot", "Nedarim", "Nazir", "Sotah", "Gittin", "Kiddushin"],
    "Nezikin":  ["Bava Kamma", "Bava Metzia", "Bava Batra", "Sanhedrin", "Makkot",
                 "Shevuot", "Eduyot", "Avodah Zarah", "Avot", "Horayot"],
    "Kodashim": ["Zevachim", "Menachot", "Chullin", "Bekhorot", "Arakhin",
                 "Temurah", "Keritot", "Meilah", "Tamid", "Middot", "Kinnim"],
    "Taharot":  ["Keilim", "Oholot", "Negaim", "Parah", "Taharot", "Mikvaot",
                 "Niddah", "Makhshirin", "Zavim", "Tevul Yom", "Yadayim", "Uktzim"],
}

TRACTATE_TO_SEDER = {t: s for s, tracts in SEDARIM.items() for t in tracts}


# ---------------------------------------------------------------------------
# Sefaria fetch
# ---------------------------------------------------------------------------

def sefaria_url(tractate: str) -> str:
    ref = urllib.parse.quote(f"Mishnah {tractate}")
    return f"https://www.sefaria.org/api/texts/{ref}?lang=en&pad=0&commentary=0"


def fetch_tractate(tractate: str) -> str:
    """Fetch English Mishnah text for a tractate from Sefaria. Returns flat readable string."""
    cache_file = DIR / "sefaria_cache" / f"{tractate.replace(' ', '_')}.json"
    cache_file.parent.mkdir(exist_ok=True)

    if cache_file.exists():
        raw = json.loads(cache_file.read_text(encoding="utf-8"))
    else:
        url = sefaria_url(tractate)
        print(f"  Fetching {url}", flush=True)
        try:
            with urllib.request.urlopen(url, timeout=30) as resp:
                raw = json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            sys.exit(f"Sefaria fetch failed for {tractate}: {e}")
        cache_file.write_text(json.dumps(raw, ensure_ascii=False), encoding="utf-8")
        time.sleep(0.5)  # polite rate limit

    chapters = raw.get("text", [])
    lines = []
    for ch_idx, chapter in enumerate(chapters, 1):
        if isinstance(chapter, list):
            for m_idx, mishna in enumerate(chapter, 1):
                text = mishna if isinstance(mishna, str) else ""
                # Strip HTML tags
                text = re.sub(r"<[^>]+>", "", text).strip()
                if text:
                    lines.append(f"{ch_idx}:{m_idx} {text}")
        elif isinstance(chapter, str):
            text = re.sub(r"<[^>]+>", "", chapter).strip()
            if text:
                lines.append(f"{ch_idx} {text}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Taxonomy helpers
# ---------------------------------------------------------------------------

def load_taxonomy() -> list[dict]:
    taxonomy_file = DIR / "merged_taxonomy.json"
    if not taxonomy_file.exists():
        taxonomy_file = DIR / "seed_taxonomy.json"
    data = json.loads(taxonomy_file.read_text(encoding="utf-8"))
    return data if isinstance(data, list) else data.get("entries", [])


def compact_taxonomy(entries: list[dict]) -> str:
    """Compact id → name list for the prompt."""
    lines = []
    for e in entries:
        parent = e.get("parent_id", "")
        cat = e.get("category", "")
        lines.append(f'{e["id"]} [{cat}]: {e["name"]}')
    return "\n".join(lines)


def make_id(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[''']", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return f"gap_{s}"


# ---------------------------------------------------------------------------
# Claude call
# ---------------------------------------------------------------------------

SYSTEM = """You are a Talmudic scholar and taxonomy expert helping build a topical index of the Talmud called Mafteach haDaf.

Your task: given the English text of a Mishnah tractate and an existing taxonomy, identify halakhic and conceptual topics that appear in this tractate but are NOT adequately represented in the existing taxonomy.

Rules:
- Only propose entries for topics with genuine independent standing in halakhic discourse — not every detail or sub-case
- Each entry needs a parent_id: either an existing taxonomy ID or a new proposed parent (prefix new parent IDs with "NEW:")
- Be specific: "Tumah of Corpse (Tum'at Met)" is better than "Ritual Impurity"
- Include Hebrew name where you know it
- Do NOT re-propose entries already in the taxonomy (check carefully)
- Return ONLY a JSON array, no explanation, no markdown fences

Entry format:
{
  "id": "gap_snake_case_name",
  "name": "English name",
  "hebrew": "עברית",
  "parent_id": "existing_taxonomy_id_or_NEW:proposed_parent_name",
  "category": "CATEGORY NAME IN CAPS",
  "sources": ["mishnah"],
  "notes": "brief note on what this covers"
}"""


def extract_candidates(tractate: str, seder: str, text: str, taxonomy_compact: str) -> list[dict]:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("ANTHROPIC_API_KEY not set")

    client = anthropic.Anthropic(api_key=api_key)

    prompt = f"""EXISTING TAXONOMY ({len(taxonomy_compact.splitlines())} entries):
{taxonomy_compact}

---
TRACTATE: Mishnah {tractate} (Seder {seder})

{text}

---
Return a JSON array of NEW taxonomy entries for topics in this tractate not yet in the taxonomy.
If there are no gaps, return an empty array: []"""

    print(f"  Sending to Claude ({len(prompt):,} chars)...", end=" ", flush=True)

    full = ""
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=8000,
        system=SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        for chunk in stream.text_stream:
            full += chunk
    print(f"received {len(full):,} chars")

    # Strip markdown fences if present
    full = full.strip()
    if full.startswith("```"):
        full = re.sub(r"^```[a-z]*\n?", "", full)
        full = re.sub(r"\n?```$", "", full)

    try:
        candidates = json.loads(full)
        if not isinstance(candidates, list):
            candidates = []
    except json.JSONDecodeError as e:
        print(f"  JSON parse error: {e} — saving raw")
        (DIR / f"gap_{tractate.replace(' ', '_')}.raw.txt").write_text(full, encoding="utf-8")
        candidates = []

    return candidates


# ---------------------------------------------------------------------------
# Process tractates
# ---------------------------------------------------------------------------

def process_tractate(tractate: str):
    seder = TRACTATE_TO_SEDER.get(tractate, "Unknown")
    out_file = DIR / f"gap_{tractate.replace(' ', '_')}.json"

    if out_file.exists():
        print(f"  {tractate}: already done ({out_file.name}), skipping. Delete to re-run.")
        return

    print(f"\n[{seder}] {tractate}")
    text = fetch_tractate(tractate)
    print(f"  Text: {len(text):,} chars, {len(text.splitlines())} mishnayot")

    taxonomy = load_taxonomy()
    taxonomy_compact = compact_taxonomy(taxonomy)

    candidates = extract_candidates(tractate, seder, text, taxonomy_compact)
    print(f"  Candidates: {len(candidates)}")

    out_file.write_text(json.dumps(candidates, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  Saved → {out_file.name}")


# ---------------------------------------------------------------------------
# Merge all gap files → candidates.json + xlsx
# ---------------------------------------------------------------------------

def merge_candidates():
    all_candidates = []
    tractate_map = {}  # id → tractate name for provenance

    gap_files = sorted(DIR.glob("gap_*.json"))
    gap_files = [f for f in gap_files if f.name != "gap_candidates.json" and not f.name.endswith(".raw.txt")]

    for f in gap_files:
        tractate = f.stem[4:].replace("_", " ")
        entries = json.loads(f.read_text(encoding="utf-8"))
        for e in entries:
            tractate_map[e.get("id", "")] = tractate
        all_candidates.extend(entries)

    # Deduplicate by id
    seen = {}
    deduped = []
    for e in all_candidates:
        eid = e.get("id", make_id(e.get("name", "")))
        if eid not in seen:
            seen[eid] = True
            deduped.append(e)

    print(f"\nMerge: {len(all_candidates)} total → {len(deduped)} after dedup")

    out_json = DIR / "gap_candidates.json"
    out_json.write_text(json.dumps({"candidates": deduped}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out_json}")

    # Build id → name lookup from merged taxonomy for parent name resolution
    id_to_name = {}
    for taxonomy_file in [DIR / "merged_taxonomy.json", DIR / "seed_taxonomy.json"]:
        if taxonomy_file.exists():
            tax_data = json.loads(taxonomy_file.read_text(encoding="utf-8"))
            tax_entries = tax_data if isinstance(tax_data, list) else tax_data.get("entries", [])
            id_to_name = {e["id"]: e["name"] for e in tax_entries}
            break

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Candidates"

    headers = ["Tractate", "Proposed ID", "Name", "Hebrew", "Parent Entry", "Category", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 40
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G1"

    seder_colors = {
        "Zeraim": "FFF2CC", "Moed": "D9E8FF", "Nashim": "FFD9F0",
        "Nezikin": "D9FFE8", "Kodashim": "FFE6CC", "Taharot": "F0D9FF",
    }

    for row_idx, e in enumerate(deduped, 2):
        tractate = tractate_map.get(e.get("id", ""), "")
        seder = TRACTATE_TO_SEDER.get(tractate, "")
        color = seder_colors.get(seder, "F0F0F0")
        vals = [
            tractate,
            e.get("id", ""),
            e.get("name", ""),
            e.get("hebrew", ""),
            id_to_name.get(e.get("parent_id", ""), e.get("parent_id", "")),
            e.get("category", ""),
            e.get("notes", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=color)
            if col == 4:
                cell.alignment = Alignment(horizontal="right")

    out_xlsx = DIR / "gap_candidates.xlsx"
    wb.save(out_xlsx)
    print(f"Wrote {out_xlsx}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Extract missing taxonomy entries from Mishnah via Sefaria")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--tractate", metavar="NAME", help="Single tractate, e.g. 'Shabbat'")
    group.add_argument("--seder", metavar="NAME", help="All tractates in a seder, e.g. 'Moed'")
    group.add_argument("--all", action="store_true", help="All 6 sedarim")
    group.add_argument("--resume", action="store_true", help="Resume interrupted --all run (skips completed tractates)")
    group.add_argument("--merge", action="store_true", help="Merge all gap_*.json → candidates.json + xlsx")
    args = parser.parse_args()

    if args.merge:
        merge_candidates()
        return

    if args.tractate:
        matched = next((t for s in SEDARIM.values() for t in s
                        if t.lower() == args.tractate.lower()), None)
        if not matched:
            sys.exit(f"Unknown tractate: {args.tractate}. Check spelling.")
        process_tractate(matched)

    elif args.seder:
        matched = next((s for s in SEDARIM if s.lower() == args.seder.lower()), None)
        if not matched:
            sys.exit(f"Unknown seder: {args.seder}. Options: {', '.join(SEDARIM)}")
        for tractate in SEDARIM[matched]:
            process_tractate(tractate)

    elif args.all or args.resume:
        if args.resume:
            done = [f.stem[4:].replace("_", " ") for f in DIR.glob("gap_*.json")
                    if not f.name.endswith(".raw.txt")]
            remaining = [(s, t) for s, tracts in SEDARIM.items() for t in tracts if t not in done]
            if not remaining:
                print("Nothing to resume — all tractates already completed. Run --merge next.")
                return
            print(f"Resuming: {len(done)} done, {len(remaining)} remaining")
        for seder, tractates in SEDARIM.items():
            print(f"\n=== Seder {seder} ===")
            for tractate in tractates:
                process_tractate(tractate)


if __name__ == "__main__":
    main()
