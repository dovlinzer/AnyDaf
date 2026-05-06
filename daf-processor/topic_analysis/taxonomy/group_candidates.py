#!/usr/bin/env python3
"""
group_candidates.py — Ask Claude to group overly-granular gap candidates under parent entries.

Processes gap_candidates.json category by category. For each category, Claude reviews
the candidates alongside the existing taxonomy entries and:
  - Creates new parent entries for clusters of 3+ related candidates
  - Reassigns parent_id for grouped entries
  - Leaves standalone entries unchanged

Output: gap_candidates_grouped.json + gap_candidates_grouped.xlsx

Usage:
  python group_candidates.py
  python group_candidates.py --category "CIVIL AND CRIMINAL LAW"   # single category
"""

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv(Path(__file__).parent.parent.parent / ".env", override=True)

DIR = Path(__file__).parent

SYSTEM = """You are a Talmudic scholar and taxonomy expert helping build Mafteach haDaf, a topical index of the Talmud.

You will receive:
1. EXISTING TAXONOMY — entries already in the taxonomy for a given category
2. GAP CANDIDATES — proposed new entries for that category, extracted from the Mishnah

Your task: reorganize the gap candidates so the taxonomy stays navigable.

Rules:
- If 3 or more candidates clearly belong to a common halakhic/conceptual grouping, create ONE parent entry for that group and nest the candidates under it
- A good parent entry has a clear, recognized name in halakhic literature (e.g. "Interaction with Gentiles", "Laws of Lost Property")
- If a suitable parent already exists in the EXISTING TAXONOMY, use that as parent_id instead of creating a new one
- For new parent entries you create, prefix the id with "group_" (e.g. "group_interaction_gentiles")
- Entries that don't belong to any cluster stay as-is with their original parent_id
- Do NOT merge distinct halakhic topics just because they're in the same tractate
- Do NOT create a parent for only 2 entries — leave them with their original parent
- Keep all candidates in the output — do not drop any

Return ONLY a JSON array containing:
- Any new parent entries you created (with parent_id pointing to an existing taxonomy entry)
- All original candidates, with parent_id updated where grouped

Entry format:
{
  "id": "entry_id",
  "name": "English name",
  "hebrew": "עברית",
  "parent_id": "existing_or_new_id",
  "category": "CATEGORY NAME",
  "sources": ["mishnah"],
  "notes": "brief note"
}"""


def load_taxonomy() -> list[dict]:
    for f in [DIR / "merged_taxonomy.json", DIR / "seed_taxonomy.json"]:
        if f.exists():
            data = json.loads(f.read_text(encoding="utf-8"))
            return data if isinstance(data, list) else data.get("entries", [])
    return []


CHUNK_SIZE = 35  # max candidates per Claude call


def call_claude(category: str, candidates: list[dict], existing: list[dict]) -> list[dict]:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        sys.exit("ANTHROPIC_API_KEY not set")

    client = anthropic.Anthropic(api_key=api_key)
    existing_compact = "\n".join(f'  {e["id"]}: {e["name"]}' for e in existing)
    candidates_json = json.dumps(candidates, ensure_ascii=False, indent=2)

    prompt = f"""CATEGORY: {category}

EXISTING TAXONOMY ENTRIES IN THIS CATEGORY ({len(existing)} entries):
{existing_compact if existing_compact else '  (none)'}

GAP CANDIDATES TO REORGANIZE ({len(candidates)} entries):
{candidates_json}

Return a JSON array with all candidates (plus any new parent entries you create).
New parent entries go BEFORE the entries nested under them."""

    print(f"  Sending {len(candidates)} candidates...", end=" ", flush=True)

    full = ""
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=16000,
        system=SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        for chunk in stream.text_stream:
            full += chunk
    print(f"received {len(full):,} chars")

    full = full.strip()
    if full.startswith("```"):
        full = re.sub(r"^```[a-z]*\n?", "", full)
        full = re.sub(r"\n?```$", "", full)

    try:
        result = json.loads(full)
        if not isinstance(result, list):
            raise ValueError("Expected JSON array")
        return result
    except (json.JSONDecodeError, ValueError) as e:
        print(f"  Parse error: {e} — keeping originals unchanged")
        return candidates


def process_category(category: str, candidates: list[dict], taxonomy_by_cat: dict) -> list[dict]:
    existing = taxonomy_by_cat.get(category, [])

    if len(candidates) <= CHUNK_SIZE:
        return call_claude(category, candidates, existing)

    # Split into chunks, process each, merge results
    results = []
    for i in range(0, len(candidates), CHUNK_SIZE):
        chunk = candidates[i:i + CHUNK_SIZE]
        print(f"  Chunk {i // CHUNK_SIZE + 1}/{(len(candidates) - 1) // CHUNK_SIZE + 1}", end=" ")
        results.extend(call_claude(category, chunk, existing))
    return results


def build_xlsx(entries: list[dict], taxonomy: list[dict], out_path: Path):
    id_to_name = {e["id"]: e["name"] for e in taxonomy + entries}

    SEDARIM_COLORS = {
        "PRAYER AND LITURGY": "D9E8FF",
        "SHABBAT": "FFF2CC", "HOLIDAYS AND FESTIVALS": "FFF2CC",
        "CIVIL AND CRIMINAL LAW": "D9FFE8",
        "FAMILY LAW": "FFD9F0",
        "TEMPLE AND SACRIFICES": "FFE6CC",
        "RITUAL PURITY AND IMPURITY": "F0D9FF",
        "AGGADA AND THEOLOGY": "FFD9D9",
        "TALMUDIC METHODOLOGY": "D9D9FF",
        "ETHICS AND CHARACTER": "D9FFD9",
        "INTERPERSONAL ETHICS": "D9FFEE",
    }

    cats = []
    seen = set()
    for e in entries:
        c = e.get("category", "")
        if c not in seen:
            cats.append(c)
            seen.add(c)

    COLORS = [
        "FFF2CC","FFE6CC","FFD9D9","FFD9F0","F0D9FF","D9E8FF","D9F5FF","D9FFE8",
        "E8FFD9","F5FFD9","D9D9FF","FFD9FF","FFF0D9","D9FFFF","E0D9FF","FFD9E0",
    ]
    cat_color = {c: COLORS[i % len(COLORS)] for i, c in enumerate(cats)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Candidates"

    headers = ["Category", "Proposed ID", "Name", "Hebrew", "Parent Entry", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    for row_idx, e in enumerate(entries, 2):
        cat = e.get("category", "")
        color = cat_color.get(cat, "F0F0F0")
        parent_name = id_to_name.get(e.get("parent_id", ""), e.get("parent_id", ""))
        is_new_parent = e.get("id", "").startswith("group_")

        vals = [cat, e.get("id", ""), e.get("name", ""), e.get("hebrew", ""), parent_name, e.get("notes", "")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=color)
            if is_new_parent:
                cell.font = Font(bold=True)
            if col == 4:
                cell.alignment = Alignment(horizontal="right")

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--category", help="Process a single category only")
    args = parser.parse_args()

    candidates_file = DIR / "gap_candidates.json"
    if not candidates_file.exists():
        sys.exit("gap_candidates.json not found — run --merge first")

    raw = json.loads(candidates_file.read_text(encoding="utf-8"))
    all_candidates = raw if isinstance(raw, list) else raw.get("candidates", [])

    taxonomy = load_taxonomy()
    taxonomy_by_cat = defaultdict(list)
    for e in taxonomy:
        taxonomy_by_cat[e.get("category", "")].append(e)

    # Group candidates by category
    by_category = defaultdict(list)
    for e in all_candidates:
        by_category[e.get("category", "Uncategorized")].append(e)

    if args.category:
        matched = next((c for c in by_category if c.lower() == args.category.lower()), None)
        if not matched:
            available = "\n".join(f"  {c}" for c in sorted(by_category))
            sys.exit(f"Category not found. Available:\n{available}")
        categories_to_process = {matched: by_category[matched]}
    else:
        categories_to_process = dict(by_category)

    print(f"Processing {len(categories_to_process)} categories, {sum(len(v) for v in categories_to_process.values())} candidates\n")

    # Check for partial progress
    progress_file = DIR / ".group_progress.json"
    progress = {}
    if progress_file.exists():
        progress = json.loads(progress_file.read_text(encoding="utf-8"))
        done = len(progress)
        print(f"Resuming: {done} categories already done")

    all_grouped = []

    for category, candidates in sorted(categories_to_process.items()):
        if category in progress:
            print(f"  [{category}] skipping (already done, {len(progress[category])} entries)")
            all_grouped.extend(progress[category])
            continue

        print(f"\n[{category}] {len(candidates)} candidates")
        grouped = process_category(category, candidates, taxonomy_by_cat)

        new_parents = [e for e in grouped if e.get("id", "").startswith("group_")]
        if new_parents:
            print(f"  Created {len(new_parents)} new parent entries: {[e['name'] for e in new_parents]}")

        progress[category] = grouped
        progress_file.write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")
        all_grouped.extend(grouped)

    # Save output
    out_json = DIR / "gap_candidates_grouped.json"
    out_json.write_text(json.dumps({"candidates": all_grouped}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nWrote {out_json} — {len(all_grouped)} entries")

    out_xlsx = DIR / "gap_candidates_grouped.xlsx"
    build_xlsx(all_grouped, taxonomy, out_xlsx)
    print(f"Wrote {out_xlsx}")

    # Clean up progress file on successful completion
    if not args.category:
        progress_file.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
